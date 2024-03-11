import win32gui, win32con, win32api
import time
from openpyxl import load_workbook
alldata = []
#输入框输入字符串
def editSetText(hwndEdit, testStr):
    #WM_SETTEXT不能用PostMessage，PostMessage参数中不能有指针。异步消息，不等待
    rst = win32gui.SendMessageTimeout(hwndEdit, win32con.WM_SETTEXT, 0, testStr, win32con.SMTO_NORMAL, 1000)
    if rst[0]==1 and rst[1]==1:
        return True
    return False

def clkcondition():
    #   EnumChildWindows 为指定的父窗口枚举子窗口
    hwndSubChildList = []
    hwndChildList = []
    hd = win32gui.GetDesktopWindow()
    win32gui.EnumChildWindows(hd, lambda hwnd, param: param.append(hwnd), hwndChildList)
    for hwnd in hwndChildList:
        if "博易大师" == win32gui.GetWindowText(hwnd):
            win32gui.EnumChildWindows(hwnd, lambda hwnd, param: param.append(hwnd), hwndSubChildList)
            for hwndinfo in hwndSubChildList:
                class_name = win32gui.GetClassName(hwndinfo)
                
                window_rect = win32gui.GetWindowRect(hwndinfo)
                x = window_rect[0]
                y = window_rect[1]
                width = window_rect[2] - x
                height = window_rect[3] - y
                if class_name == "Edit" and x == 176:
                    editSetText(hwndinfo, 'ni2405')#将'testStr'字符串设置到exclue中


                if "条件" in win32gui.GetWindowText(hwndinfo):
                    win32gui.PostMessage(hwndinfo, win32con.WM_LBUTTONDOWN, win32con.MK_LBUTTON, 0)
                    win32gui.PostMessage(hwndinfo, win32con.WM_LBUTTONUP, win32con.MK_LBUTTON, 0)



def set_high_level(hwnd, data):
    hwndSubChildList = []
    # 获取窗口标题
    # print("set_high_level: data=", data)
    title = win32gui.GetWindowText(hwnd)
    # 获取窗口类名
    class_name = win32gui.GetClassName(hwnd)
    # 输出窗口信息
    if "条件单设置" == title:
        win32gui.EnumChildWindows(hwnd, lambda hwnd, param: param.append(hwnd), hwndSubChildList)
        for hwndinfodest in hwndSubChildList:
            if win32gui.GetWindowText(hwndinfodest) == "高级":
                print("句柄: ",hwndinfodest,"标题：",win32gui.GetWindowText(hwndinfodest))
                win32api.SendMessage(hwndinfodest,win32con.WM_LBUTTONDOWN,win32con.MK_LBUTTON,0)#向窗口发送模拟鼠标点击
                win32api.SendMessage(hwndinfodest,win32con.WM_LBUTTONUP,win32con.MK_LBUTTON,0)#模拟释放鼠标左键

def set_windows_data(hwnd, data):
    # 获取窗口标题
    global alldata
    title = win32gui.GetWindowText(hwnd)
    # 获取窗口类名
    class_name = win32gui.GetClassName(hwnd)
    hwndSubChildList = []
    # 输出窗口信息
    if "条件单设置" == title:
        # print("set_windows_data: data=", data)
        # print("contract=",alldata[int(data)].contract, 
        # "num=", alldata[int(data)].num, " buyadd=", alldata[int(data)].buyadd,
        # " buyloss=", alldata[int(data)].buyloss, " buyearn=", alldata[int(data)].buyearn)
        win32gui.EnumChildWindows(hwnd, lambda hwnd, param: param.append(hwnd), hwndSubChildList)
        for hwndinfodest in hwndSubChildList:
            class_name = win32gui.GetClassName(hwndinfodest)
            window_rect = win32gui.GetWindowRect(hwndinfodest)
            x = window_rect[0]
            y = window_rect[1]
            width = window_rect[2] - x
            height = window_rect[3] - y
            if class_name == "TXPPriceEdit":
                if width==70 and height == 20:
                   
                    editSetText(hwndinfodest, str(alldata[int(data)].buyadd))#将'testStr'字符串设置到exclue中
                elif width==62 and height == 20:
                    if x==1065:
                        
                        editSetText(hwndinfodest, str(alldata[int(data)].buyearn))#将'testStr'字符串设置到exclue中
                    elif x==803:
                      
                        editSetText(hwndinfodest, str(alldata[int(data)].buyloss))#将'testStr'字符串设置到exclue中
            if class_name == "TXPEdit" and width==34 and height == 20:
               
                editSetText(hwndinfodest, str(alldata[int(data)].num))#将'testStr'字符串设置到exclue中

            
            if class_name == "Edit":
                window_rect = win32gui.GetWindowRect(hwndinfodest)
                x = window_rect[0]
                y = window_rect[1]
                width = window_rect[2] - x
                height = window_rect[3] - y
                editSetText(hwndinfodest, alldata[int(data)].contract)#将'testStr'字符串设置到exclue中

confirmbtn = None
def click_confirm_btn(hwnd, data):
    global alldata
    global confirmbtn
    hwndSubChildList = []
    
    if alldata[int(data)].isadd == 1:
        return
    # 获取窗口标题
    title = win32gui.GetWindowText(hwnd)
    
    # 获取窗口类名
    # class_name = win32gui.GetClassName(hwnd)
    # 输出窗口信息
    if "条件单设置" == title:
        win32gui.EnumChildWindows(hwnd, lambda hwnd, param: param.append(hwnd), hwndSubChildList)
        for hwndinfodest in hwndSubChildList:
            
            class_name = win32gui.GetClassName(hwndinfodest)
            window_rect = win32gui.GetWindowRect(hwndinfodest)
            x = window_rect[0]
            y = window_rect[1]
            width = window_rect[2] - x
            height = window_rect[3] - y
            if class_name == "TXPButton":
                if width==75 and height == 24 and x==1015:
                    # print("句柄: ",hwndinfodest,"标题：",win32gui.GetWindowText(hwndinfodest))
                    # print("hwndinfodest=", hwndinfodest);
                    confirmbtn = hwndinfodest
                    alldata[int(data)].isadd = 1
                   

class ExcelFileInfo:
    contract=""
    buyadd = 0
    buyloss = 0
    buyearn = 0;
    num = 0
    isadd = 0
    def __init__(self):
        self.contract=""
 
# 加载工作簿
workbook = load_workbook(filename='data.xlsx')
 
# 选择工作表
sheet = workbook.active
 
# 遍历工作表中的所有行
cnt=0

for row in sheet.iter_rows(values_only=True):
    cnt=cnt+1
    if cnt > 2:
        idx=0;
        addinfo = ExcelFileInfo();
        lossinfo = ExcelFileInfo();
        addinfo.isadd = 0;
        lossinfo.isadd = 0;
        for rowinfo in row:
            if idx == 0:
                 addinfo.contract=rowinfo
                 lossinfo.contract=rowinfo
            elif idx == 1:
                 addinfo.num=rowinfo
                 lossinfo.num=rowinfo
            elif idx == 2:
                 addinfo.buyadd=rowinfo
            elif idx == 3:
                 addinfo.buyloss=rowinfo
            elif idx == 4:
                 addinfo.buyearn=rowinfo
            elif idx == 6:
                 lossinfo.buyadd=rowinfo
            elif idx == 7:
                 lossinfo.buyloss=rowinfo
            elif idx == 8:
                 lossinfo.buyearn=rowinfo
            idx=idx+1
        alldata.append(addinfo);
        alldata.append(lossinfo)

for i in range(len(alldata)):
    clkcondition()
    time.sleep(2)
    win32gui.EnumWindows(set_high_level, "set_high_level")
    win32gui.EnumWindows(set_windows_data, str(i))
    win32gui.EnumWindows(click_confirm_btn, str(i))
    time.sleep(2)
    if  confirmbtn != None:
        win32api.SendMessage(confirmbtn,win32con.WM_LBUTTONDOWN,win32con.MK_LBUTTON,0)#向窗口发送模拟鼠标点击
        win32api.SendMessage(confirmbtn,win32con.WM_LBUTTONUP,win32con.MK_LBUTTON,0)#模拟释放鼠标左键
    time.sleep(1)
    # if i == 1:
        # break
         
    
    # print("contract=",data.contract, 
    #       "num=", data.num, " buyadd=", data.buyadd,
    #       " buyloss=", data.buyloss, " buyearn=", data.buyearn)




# 